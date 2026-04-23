"""
push_demand_plan.py  —  reads Excel files and uploads a comprehensive JSON blob to DBFS.

Sources:
    Demand Plan Dashboard.xlsx  (Pivot Raw sheet — UIF, DP/APO, Franchise data)
    OP Submit.xlsx              (OP plan metrics + Seasonal Revenue)

Run:    py push_demand_plan.py
Override: py push_demand_plan.py --dashboard-xlsx "path" --op-submit-xlsx "path"
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import subprocess
import sys
import tempfile
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-8s %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger(__name__)

DBFS_PATH          = "dbfs:/FileStore/ebp_dashboard/demand_plan_blob.json"
DATABRICKS_PROFILE = "4428761713917856"

_DEFAULT_DIR = Path(Path(__file__).parent.parent)

# ── Constants ─────────────────────────────────────────────────────────────────

GEO_NORMALIZE = {
    "NORTH AMERICA": "NA", "NORTH_AMERICA": "NA", "NA": "NA",
    "EUROPEAFRICA": "EMEA", "EUROPE AFRICA": "EMEA", "EMEA": "EMEA",
    "GREATER CHINA": "GC", "GC": "GC",
    "APLA": "APLA", "ASIA PACIFIC LATIN AMERICA": "APLA",
}

DIV_MAP = {
    "APPAREL DIVISION": "AP", "FOOTWEAR DIVISION": "FW",
    "EQUIPMENT DIVISION": "ACC",
}
# Product Engine codes only (OP / APO / SRO / Exec / UIF filters)
PRODUCT_ENGINE = ("AP", "FW", "ACC")


def pe_labels_in_data(seen: set) -> list:
    """Ordered PE labels present in data; if none, show all three so the UI is usable."""
    out = [x for x in PRODUCT_ENGINE if x in seen]
    return out or list(PRODUCT_ENGINE)

BRAND_SHORT = {"NIKE": "Nike", "JORDAN": "Jordan"}
BRAND_LONG  = {"NIKE": "Nike Brand", "JORDAN": "Jordan Brand"}

CHANNEL_MAP = {
    "Nike Direct Digital Commerce": "NDDC", "Nike Stores Owned": "NSO",
    "Nike Value Stores": "NVS", "Nike Stores Partnered": "NSP",
    "Nike Marketplace Partners": "NMP",
}

GATE_ORDER     = ["GCA", "MAR", "SCP1", "SCP2", "SCP3"]
DP_VALID_GATES = {
    "GCA", "MAR", "SCP1", "SCP2", "SCP3", "F1", "F2", "F3", "F4",
    "POST_PROD_BRIEF", "POST_PROD_ALIGN", "PRE_PROD_FINAL",
    "GEO_STYLE", "GEO_STYLE_COLOR", "GBL_GEO_PIVOT", "POST_MPU-REVIEW",
}
UIF_DPSUB_GATES = {
    "POST_PROD_BRIEF", "POST_PROD_ALIGN", "PRE_PROD_FINAL",
    "GEO_STYLE", "GEO_STYLE_COLOR", "GBL_GEO_PIVOT", "POST_MPU-REVIEW",
    "F1", "F2", "F3", "F4",
}

NIKE_PRIMARY_SPORTS = [
    "Running", "Training", "Basketball", "Global Football", "Sportswear", "Kids",
]

MAPPED_SPORT_ORDER = [
    "Running", "Training", "Basketball", "Global Football", "Sportswear", "Kids",
    "Tennis", "Golf", "ACG", "Skate", "NikeSkims", "American Football", "Baseball",
    "Specialty Sports", "Cricket", "Streetwear", "FBAT", "Other",
]

NIKE_SPORT_MAP = {
    "RUNNING": "Running", "TRAINING": "Training", "BASKETBALL": "Basketball",
    "GLOBAL FOOTBALL": "Global Football", "SPORTSWEAR": "Sportswear",
    "TENNIS": "Tennis", "GOLF": "Golf", "SKATE": "Skate",
    "AMERICAN FOOTBALL": "American Football", "BASEBALL": "Baseball",
    "SPECIALTY SPORTS": "Specialty Sports", "CRICKET": "Cricket",
}

JORDAN_SPORT_MAP = {
    "STREETWEAR": "Streetwear", "BASKETBALL": "Basketball", "GOLF": "Golf",
    "AMERICAN FOOTBALL": "American Football", "GLOBAL FOOTBALL": "Global Football",
    "BASEBALL": "Baseball", "TRAINING": "Training",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def sf(v):
    try:
        return float(v) if v is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def normalize_geo(raw):
    return GEO_NORMALIZE.get((raw or "").strip().upper(), (raw or "").strip())


def normalize_brand_short(raw):
    u = (raw or "").strip().upper()
    if "JORDAN" in u: return "Jordan"
    if "NIKE" in u:   return "Nike"
    if u in ("", "NOT_SUPPLD", "UNKNOWN", "NONE", "NAN"):
        return "Nike"
    return None


def normalize_brand_long(raw):
    u = (raw or "").strip().upper()
    if "JORDAN" in u: return "Jordan Brand"
    if "NIKE" in u:   return "Nike Brand"
    if u in ("", "NOT_SUPPLD", "UNKNOWN", "NONE", "NAN"):
        return "Nike Brand"
    return None


def normalize_gate(raw):
    g = (raw or "").strip().upper().replace(" ", "").replace("/", "")
    if g in GATE_ORDER or g == "LY":
        return g
    return None


def normalize_season(raw):
    s = str(raw).strip()
    if len(s) >= 6 and s[:4].isdigit() and s[4:].isalpha():
        return s[4:].upper() + s[:4]
    return s


def season_sort_key(s):
    p = {"FA": 0, "HO": 1, "SP": 2, "SU": 3}
    if s and len(s) >= 6 and s[:2].upper() in p:
        return (s[2:], p[s[:2].upper()])
    return (s, 9)


def _normalize_consumer(raw):
    u = (raw or "").strip().upper()
    if u in ("MENS", "MEN"):     return "Mens"
    if u in ("WOMENS", "WOMEN"): return "Womens"
    if u in ("KIDS", "KID"):     return "Kids"
    return "Other"


# ── Sport mapping — UIF filter options ────────────────────────────────────────

def nike_sport(gsf, con, sb):
    con_u = (con or "").strip().upper()
    if con_u in ("KIDS", "KID"):
        return "Kids"
    if con_u not in ("MENS", "WOMENS", "MEN", "WOMEN"):
        return None
    sb_u = (sb or "").strip().upper()
    if "ACG" in sb_u:
        return "ACG"
    if "SKIMS" in sb_u or sb_u == "NIKESKIMS":
        return "NikeSkims"
    gsf_u = (gsf or "").strip().upper()
    if gsf_u == "STREETWEAR":
        return None
    return NIKE_SPORT_MAP.get(gsf_u) or ((gsf or "").strip() or None)


def jordan_sport(gsf, con):
    con_u = (con or "").strip().upper()
    if con_u in ("KIDS", "KID"):
        return "Kids"
    if con_u not in ("MENS", "WOMENS", "MEN", "WOMEN"):
        return None
    gsf_u = (gsf or "").strip().upper()
    if gsf_u == "STREETWEAR": return "Streetwear"
    if gsf_u == "BASKETBALL": return "Basketball"
    if gsf_u == "GOLF":       return "Golf"
    return "FBAT"


# ── Sport mapping — OP / DP records (OPDP tab) ──────────────────────────────

def _map_sport_jordan(consumer, sport):
    con = (consumer or "").upper()
    if con in ("KIDS", "KID"):
        return "Kids"
    return JORDAN_SPORT_MAP.get((sport or "").upper())


def _map_sport_op(consumer, sub_brand, sport, brand=""):
    if "JORDAN" in (brand or "").upper():
        return _map_sport_jordan(consumer, sport)
    sb  = (sub_brand or "").upper()
    sp  = (sport or "").upper()
    con = (consumer or "").upper()
    if con in ("KIDS", "KID"):
        return "Kids"
    is_mw = con in ("MENS", "WOMENS")
    if is_mw and "ACG" in sb:
        return "ACG"
    if is_mw and ("SKIMS" in sb or sb == "NIKESKIMS"):
        return "NikeSkims"
    if not is_mw:
        return "Other"
    return NIKE_SPORT_MAP.get(sp, "Other")


def _map_sport_dp(consumer, sub_brand, sport, brand=""):
    return _map_sport_op(consumer, sub_brand, sport, brand)


# ── Sport mapping — SRO records ──────────────────────────────────────────────

def _map_sport_sro(consumer, sub_brand, sport, brand=""):
    if "JORDAN" in (brand or "").upper():
        con = (consumer or "").upper()
        if con in ("KIDS", "KID"):
            return "Kids"
        return JORDAN_SPORT_MAP.get((sport or "").upper()) or "Other"
    con = (consumer or "").upper()
    sb  = (sub_brand or "").upper()
    sp  = (sport or "").upper()
    if con in ("KIDS", "KID"):
        return "Kids"
    is_mw = con in ("MENS", "WOMENS")
    if is_mw and "ACG" in sb:
        return "ACG"
    if is_mw and ("SKIMS" in sb or sb == "NIKESKIMS"):
        return "NikeSkims"
    if not is_mw:
        return "Other"
    return NIKE_SPORT_MAP.get(sp, "Other")


# ── Sport mapping — Franchise (DPFR) records ─────────────────────────────────

def _map_sport_dpfr(consumer, sub_brand, sport, brand=""):
    if "JORDAN" in (brand or "").upper():
        con = (consumer or "").upper()
        if con in ("KIDS", "KID"):
            return "Kids"
        return JORDAN_SPORT_MAP.get((sport or "").upper()) or "Other"
    con = (consumer or "").upper()
    sb  = (sub_brand or "").upper()
    sp  = (sport or "").upper()
    if con in ("KIDS", "KID"):
        return "Kids"
    is_mw = con in ("MENS", "WOMENS")
    if is_mw and "ACG" in sb:
        return "ACG"
    if is_mw and ("SKIMS" in sb or sb == "NIKESKIMS"):
        return "NikeSkims"
    if not is_mw:
        return "Other"
    return NIKE_SPORT_MAP.get(sp, "Other")


# ══════════════════════════════════════════════════════════════════════════════
# Data loading — Demand Plan Dashboard.xlsx  (single pass → UIF + DP + FR)
# ══════════════════════════════════════════════════════════════════════════════

def load_dashboard_data(path: Path):
    """Single pass through Pivot Raw → (uif_rows, dp_rows, fr_rows)."""
    logger.info("Reading Demand Plan Dashboard: %s", path)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb["Pivot Raw"]

    uif_agg = defaultdict(lambda: [0.0, 0.0])
    dp_agg  = defaultdict(lambda: [0.0, 0.0])
    fr_agg  = defaultdict(lambda: [0.0, 0.0])
    n = 0

    for row in ws.iter_rows(min_row=5, values_only=True):
        v = list(row)
        n += 1

        season_raw = str(v[4]).strip() if len(v) > 4 and v[4] else ""
        season = normalize_season(season_raw) if season_raw else ""
        if not season or season.upper() in ("TOTAL", "GRAND", "NAN"):
            continue

        brand_raw = str(v[2]).strip().upper() if len(v) > 2 and v[2] else ""
        brand_s = BRAND_SHORT.get(brand_raw) or normalize_brand_short(brand_raw)
        brand_l = BRAND_LONG.get(brand_raw) or normalize_brand_long(brand_raw)
        if not brand_s or not brand_l:
            continue

        geo       = normalize_geo(str(v[0]).strip() if v[0] else "")
        div_raw   = str(v[1]).strip() if len(v) > 1 and v[1] else ""
        div_short = DIV_MAP.get(div_raw, "")
        consumer  = str(v[6]).strip() if len(v) > 6 and v[6] else ""
        gsf       = str(v[7]).strip() if len(v) > 7 and v[7] else ""
        sub_brand = str(v[3]).strip() if len(v) > 3 and v[3] else ""
        fr_raw    = str(v[9]).strip() if len(v) > 9 and v[9] else ""
        mc        = str(v[10]).strip() if len(v) > 10 and v[10] else ""
        pf        = str(v[11]).strip() if len(v) > 11 and v[11] else ""
        psf       = str(v[12]).strip() if len(v) > 12 and v[12] else ""
        bnb       = str(v[13]).strip() if len(v) > 13 and v[13] else ""

        whs    = sf(v[14] if len(v) > 14 else None)
        dp_qty = sf(v[15] if len(v) > 15 else None)

        gate_code = str(v[16]).strip().upper() if len(v) > 16 and v[16] else ""
        gate_col5 = str(v[5]).strip() if len(v) > 5 and v[5] else ""

        # ── UIF aggregation (Code Gate = col Q) ──
        gate_norm = normalize_gate(gate_code)
        if gate_norm:
            uif_key = (season, gate_norm, brand_s, geo, div_raw, consumer, gsf,
                       sub_brand, fr_raw, pf, psf, bnb, mc)
            a = uif_agg[uif_key]
            a[0] += dp_qty
            a[1] += whs

        # ── UIF aggregation for DP Submit gates (col F) ──
        if gate_col5 in UIF_DPSUB_GATES:
            uif_dp_key = (season, gate_col5, brand_s, geo, div_raw, consumer, gsf,
                          sub_brand, fr_raw, pf, psf, bnb, mc)
            a = uif_agg[uif_dp_key]
            a[0] += dp_qty
            a[1] += whs

        # ── DP / APO aggregation (Gate = col F) ──
        if gate_col5 in DP_VALID_GATES:
            msp = _map_sport_dp(consumer, sub_brand, gsf, brand_l)
            if msp:
                dp_key = (season, gate_col5, brand_l, geo, div_short,
                          sub_brand, gsf, msp, pf, psf, consumer)
                a = dp_agg[dp_key]
                a[0] += dp_qty
                a[1] += whs

        # ── Franchise aggregation (Gate = col F) ──
        if gate_col5:
            msp_fr = _map_sport_dpfr(consumer, sub_brand, gsf, brand_l)
            if msp_fr is not None:
                fr_up = fr_raw.upper()
                franchise = ("_Other" if (not fr_raw or fr_up in
                             ("NO FRANCHISE", "NOT_SUPPLD", "NONE", "NAN"))
                             else fr_raw)
                sb_n = sub_brand.strip() if sub_brand else ""
                sb_n = ("Not Supplied" if (not sb_n or sb_n.upper() in
                        ("NOT_SUPPLD", "NONE", "NAN", ""))
                        else sb_n)
                fr_key = (season, gate_col5, geo, div_short, brand_l,
                          consumer, msp_fr, franchise, pf, psf, sb_n, bnb)
                a = fr_agg[fr_key]
                a[0] += dp_qty
                a[1] += whs

    wb.close()

    uif = [
        {"s": s, "g": g, "b": b, "geo": geo, "div": div,
         "d": DIV_MAP.get(str(div).strip(), ""),
         "con": con, "gsf": gsf, "sb": sb, "fr": fr, "pf": pf, "psf": psf, "bnb": bnb, "mc": mc,
         "q": round(vals[0], 2), "w": round(vals[1], 2)}
        for (s, g, b, geo, div, con, gsf, sb, fr, pf, psf, bnb, mc), vals
        in uif_agg.items()
    ]

    dp = [
        {"s": s, "gt": gt, "b": b, "g": g, "d": d, "sb": sb, "sp": sp,
         "msp": msp, "pf": pf, "psf": psf, "con": con,
         "adp": round(vals[0], 2), "afpd": round(vals[1], 2)}
        for (s, gt, b, g, d, sb, sp, msp, pf, psf, con), vals
        in dp_agg.items()
    ]

    fr = [
        {"s": s, "gt": gt, "geo": geo, "dv": dv, "b": b, "con": con,
         "sp": sp, "fr": fr_, "pf": pf, "psf": psf, "sb": sb, "bnb": bnb,
         "u": round(vals[0], 0), "fpd": round(vals[1], 2)}
        for (s, gt, geo, dv, b, con, sp, fr_, pf, psf, sb, bnb), vals
        in fr_agg.items()
    ]

    logger.info("  %d raw rows → UIF: %d, DP: %d, FR: %d",
                n, len(uif), len(dp), len(fr))
    return uif, dp, fr


# ══════════════════════════════════════════════════════════════════════════════
# Data loading — OP Submit.xlsx  (single pass → OP + SRO)
# ══════════════════════════════════════════════════════════════════════════════

def load_op_data(path: Path):
    """Read OP Submit → (op_records, sro_records) with mapped sport."""
    logger.info("Reading OP Submit: %s", path)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    op_agg  = defaultdict(lambda: [0.0] * 12)
    sro_agg = defaultdict(lambda: [0.0, 0.0])
    n = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        v = list(row)
        if len(v) < 10:
            continue
        n += 1

        season = normalize_season(str(v[8]).strip()) if v[8] else ""
        if not season:
            continue

        brand_raw = str(v[2]).strip().upper() if v[2] else ""
        brand = BRAND_LONG.get(brand_raw) or normalize_brand_long(brand_raw)
        if not brand:
            continue

        consumer_raw = str(v[3]).strip() if len(v) > 3 and v[3] else ""
        geo      = normalize_geo(str(v[0]).strip() if v[0] else "")
        div_raw  = str(v[4]).strip() if len(v) > 4 and v[4] else ""
        div      = DIV_MAP.get(div_raw, "")
        sb       = str(v[5]).strip() if len(v) > 5 and v[5] else ""
        sport    = str(v[6]).strip() if len(v) > 6 and v[6] else ""
        chan_raw  = str(v[1]).strip() if v[1] else ""
        chan      = CHANNEL_MAP.get(chan_raw, "")

        # ── OP record ──
        msp_op = _map_sport_op(consumer_raw, sb, sport, brand)
        if msp_op:
            con = _normalize_consumer(consumer_raw)
            op_key = (season, brand, geo, div, sb, sport, chan, msp_op, con)
            a = op_agg[op_key]
            a[0]  += sf(v[15] if len(v) > 15 else None)   # sr      (col P)
            a[1]  += sf(v[9]  if len(v) > 9  else None)   # rsv     (col J)
            a[2]  += sf(v[21] if len(v) > 21 else None)   # dp      (GF+AA+AO units, col V)
            a[3]  += sf(v[18] if len(v) > 18 else None)   # fpd     (GF+AA+AO $,     col S)
            a[4]  += sf(v[16] if len(v) > 16 else None)   # sr_py   (col Q)
            a[5]  += sf(v[10] if len(v) > 10 else None)   # rsv_py  (col K)
            a[6]  += sf(v[22] if len(v) > 22 else None)   # dp_py   (GF+AA+AO units PY, col W)
            a[7]  += sf(v[19] if len(v) > 19 else None)   # fpd_py  (GF+AA+AO $ PY,     col T)
            a[8]  += sf(v[24] if len(v) > 24 else None)   # sru     (col Y)
            a[9]  += sf(v[25] if len(v) > 25 else None)   # sru_py  (col Z)
            a[10] += sf(v[12] if len(v) > 12 else None)   # rsu     (col M)
            a[11] += sf(v[13] if len(v) > 13 else None)   # rsu_py  (col N)

        # ── SRO record ──
        msp_sro = _map_sport_sro(consumer_raw, sb, sport, brand)
        if msp_sro:
            con_upper = consumer_raw.upper()
            sro_key = (season, brand, con_upper, geo, div, sb, sport, chan, msp_sro)
            a = sro_agg[sro_key]
            a[0] += sf(v[15] if len(v) > 15 else None)   # sr TY
            a[1] += sf(v[16] if len(v) > 16 else None)   # sr PY

    wb.close()

    op = [
        {"s": s, "b": b, "g": g, "d": d, "sb": sb, "sp": sp, "ch": ch,
         "msp": msp, "con": con,
         "sr": round(vals[0], 2), "rsv": round(vals[1], 2),
         "dp": round(vals[2], 2), "fpd": round(vals[3], 2),
         "srP": round(vals[4], 2), "rsvP": round(vals[5], 2),
         "dpP": round(vals[6], 2), "fpdP": round(vals[7], 2),
         "sru": round(vals[8], 2), "sruP": round(vals[9], 2),
         "rsu": round(vals[10], 2), "rsuP": round(vals[11], 2)}
        for (s, b, g, d, sb, sp, ch, msp, con), vals in op_agg.items()
    ]

    sro = [
        {"s": s, "b": b, "con": con, "g": g, "d": d, "sb": sb,
         "sp": sp, "ch": ch, "msp": msp,
         "sr": round(vals[0], 2), "srP": round(vals[1], 2)}
        for (s, b, con, g, d, sb, sp, ch, msp), vals in sro_agg.items()
    ]

    logger.info("  %d raw rows → OP: %d, SRO: %d", n, len(op), len(sro))
    return op, sro


# ══════════════════════════════════════════════════════════════════════════════
# Filter option builders
# ══════════════════════════════════════════════════════════════════════════════

def get_uif_filter_options(rows):
    opts = {"Nike": {}, "Jordan": {}}
    for brand in ("Nike", "Jordan"):
        br = [r for r in rows if r["b"] == brand]
        opts[brand]["geo"]       = sorted({r["geo"] for r in br if r["geo"]})
        opts[brand]["season"]    = sorted({r["s"]   for r in br if r["s"]},
                                          key=season_sort_key, reverse=True)
        pe_seen = {r.get("d") for r in br if r.get("d")}
        opts[brand]["division"] = pe_labels_in_data(pe_seen)
        opts[brand]["sub_brand"] = sorted({r["sb"]  for r in br if r["sb"]})
        opts[brand]["pf"]        = sorted({r["pf"]  for r in br if r["pf"]})
        opts[brand]["psf"]       = sorted({r["psf"] for r in br if r["psf"]})
        opts[brand]["bnb"]       = sorted({r["bnb"] for r in br if r["bnb"]})
        if brand == "Nike":
            sp_set = set()
            for r in br:
                sp = nike_sport(r["gsf"], r["con"], r["sb"])
                if sp:
                    sp_set.add(sp)
            primary = [s for s in NIKE_PRIMARY_SPORTS if s in sp_set]
            detail  = sorted(sp_set - set(NIKE_PRIMARY_SPORTS))
            opts[brand]["sport"] = primary + detail
        else:
            opts[brand]["sport"] = ["Streetwear", "Basketball", "Golf", "FBAT", "Kids"]
    return opts


def get_opdp_filter_options(op_recs, dp_recs):
    seasons = set(); brands = set(); geos = set(); divs = set()
    sub_brands = set(); sports = set(); consumers = set()
    pdd_families = set(); pdd_sub_families = set()

    for r in op_recs:
        if r.get("s"):   seasons.add(r["s"])
        if r.get("b"):   brands.add(r["b"])
        if r.get("g"):   geos.add(r["g"])
        if r.get("d"):   divs.add(r["d"])
        if r.get("sb"):  sub_brands.add(r["sb"])
        if r.get("msp"): sports.add(r["msp"])
        if r.get("con"): consumers.add(r["con"])
    for r in dp_recs:
        if r.get("s"):   seasons.add(r["s"])
        if r.get("b"):   brands.add(r["b"])
        if r.get("g"):   geos.add(r["g"])
        if r.get("d"):   divs.add(r["d"])
        if r.get("sb"):  sub_brands.add(r["sb"])
        if r.get("msp"): sports.add(r["msp"])
        if r.get("pf"):  pdd_families.add(r["pf"])
        if r.get("psf"): pdd_sub_families.add(r["psf"])
        if r.get("con"): consumers.add(r["con"])

    fbat_set = {"American Football", "Global Football", "Baseball", "Training"}
    if sports & fbat_set:
        sports.add("FBAT")

    sp_order  = {s: i for i, s in enumerate(MAPPED_SPORT_ORDER)}
    con_order = {"Mens": 0, "Womens": 1, "Kids": 2, "Other": 3}
    return {
        "seasons":          sorted(seasons, key=season_sort_key, reverse=True),
        "brands":           sorted(brands),
        "geos":             sorted(geos),
        "divs":             pe_labels_in_data(divs),
        "sub_brands":       sorted(sub_brands),
        "sports":           sorted(sports, key=lambda x: sp_order.get(x, 99)),
        "consumers":        sorted(consumers, key=lambda x: con_order.get(x, 99)),
        "pdd_families":     sorted(pdd_families),
        "pdd_sub_families": sorted(pdd_sub_families),
    }


def get_sro_filter_options(recs):
    seasons = set(); brands = set(); geos = set(); divs = set()
    sub_brands = set(); sports = set()
    for r in recs:
        if r["s"]:   seasons.add(r["s"])
        if r["b"]:   brands.add(r["b"])
        if r["g"]:   geos.add(r["g"])
        if r["d"]:   divs.add(r["d"])
        if r["sb"]:  sub_brands.add(r["sb"])
        if r["msp"]: sports.add(r["msp"])
    fbat_set = {"American Football", "Global Football", "Baseball", "Training"}
    if sports & fbat_set:
        sports.add("FBAT")
    sp_order = {s: i for i, s in enumerate(MAPPED_SPORT_ORDER)}
    return {
        "seasons":    sorted(seasons, key=season_sort_key),
        "brands":     sorted(brands),
        "geos":       sorted(geos),
        "divs":       pe_labels_in_data(divs),
        "sub_brands": sorted(sub_brands),
        "sports":     sorted(sports, key=lambda x: sp_order.get(x, 99)),
    }


def get_dpfr_filter_options(recs):
    seasons = set(); geos = set(); divs = set(); brands = set()
    consumers = set(); sports = set(); sub_brands = set(); bnbs = set()
    pdd_families = set(); pdd_sub_families = set(); franchises = set()
    for r in recs:
        if r.get("s"):   seasons.add(r["s"])
        if r.get("geo"): geos.add(r["geo"])
        if r.get("dv"):  divs.add(r["dv"])
        if r.get("b"):   brands.add(r["b"])
        if r.get("con"): consumers.add(r["con"])
        if r.get("sp"):  sports.add(r["sp"])
        if r.get("sb"):  sub_brands.add(r["sb"])
        if r.get("pf"):  pdd_families.add(r["pf"])
        if r.get("psf"): pdd_sub_families.add(r["psf"])
        if r.get("bnb"): bnbs.add(r["bnb"])
        if r.get("fr"):  franchises.add(r["fr"])
    fbat_set = {"American Football", "Global Football", "Baseball", "Training"}
    if sports & fbat_set:
        sports.add("FBAT")
    return {
        "seasons":          sorted(seasons, key=season_sort_key),
        "geos":             sorted(geos),
        "divs":             pe_labels_in_data(divs),
        "brands":           sorted(brands),
        "sub_brands":       sorted(sub_brands),
        "consumers":        sorted(consumers),
        "sports":           sorted(sports),
        "pdd_families":     sorted(pdd_families),
        "pdd_sub_families": sorted(pdd_sub_families),
        "bnbs":             sorted(bnbs),
        "franchises":       sorted(franchises),
    }


def default_season_from_op_dp(op_records: list, dp_rows: list):
    """Default exec season: newest in APO/DP (``dp``) when present, else newest OP-only.

    See ``refresh_from_snowflake.default_season_from_op_dp`` for rationale
    (avoid preselecting a plan-only future season with no APO rows).
    """
    op_codes = {r.get("s") for r in op_records if r.get("s")}
    dp_codes = {r.get("s") for r in dp_rows if r.get("s")}
    if not op_codes and not dp_codes:
        return None
    sord = {"FA": 0, "HO": 1, "SP": 2, "SU": 3}

    def _key(s: str):
        s = (s or "").strip()
        if len(s) < 3:
            return (0, 99)
        pre = s[:2].upper()
        try:
            y = int(s[2:])
        except ValueError:
            y = 0
        return (y, sord.get(pre, 99))

    if dp_codes:
        return max(dp_codes, key=_key)
    return max(op_codes, key=_key) if op_codes else None


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════

def parse_args():
    p = argparse.ArgumentParser(description="Build and upload dashboard JSON blob to DBFS")
    p.add_argument("--dashboard-xlsx", type=Path,
                   default=_DEFAULT_DIR / "Demand Plan Dashboard.xlsx")
    p.add_argument("--op-submit-xlsx", type=Path,
                   default=_DEFAULT_DIR / "OP Submit.xlsx")
    p.add_argument("--dbfs-path", default=DBFS_PATH)
    p.add_argument("--local-out", type=Path, default=None,
                   help="Write blob to local file instead of uploading to DBFS")
    return p.parse_args()


def main():
    import time
    args = parse_args()
    t0 = time.time()

    dashboard_xlsx = args.dashboard_xlsx.resolve()
    op_submit_xlsx = args.op_submit_xlsx.resolve()
    dbfs_path      = args.dbfs_path

    for p in [dashboard_xlsx, op_submit_xlsx]:
        if not p.exists():
            logger.error("File not found: %s", p)
            sys.exit(1)

    logger.info("=== push_demand_plan.py ===")
    logger.info("  Dashboard : %s", dashboard_xlsx)
    logger.info("  OP Submit : %s", op_submit_xlsx)
    logger.info("  DBFS path : %s", dbfs_path)

    uif_rows, dp_rows, fr_rows = load_dashboard_data(dashboard_xlsx)
    op_records, sro_records     = load_op_data(op_submit_xlsx)

    ds = default_season_from_op_dp(op_records, dp_rows)
    blob = {
        "m": {
            "ts":              datetime.now(timezone.utc).isoformat(),
            "gates":           GATE_ORDER,
            "src":             "excel-demand-plan",
            "dashboard_path":  str(dashboard_xlsx),
            "op_submit_path":   str(op_submit_xlsx),
            "default_season":  ds,
        },
        "u":   uif_rows,
        "uo":  get_uif_filter_options(uif_rows),
        "op":  op_records,
        "dp":  dp_rows,
        "oo":  get_opdp_filter_options(op_records, dp_rows),
        "sro": sro_records,
        "so":  get_sro_filter_options(sro_records),
        "fr":  fr_rows,
        "fo":  get_dpfr_filter_options(fr_rows),
    }

    blob_json = json.dumps(blob, separators=(",", ":"))
    size_mb   = len(blob_json) / (1024 * 1024)
    logger.info("Blob size : %.2f MB", size_mb)

    if args.local_out:
        local_path = args.local_out.resolve()
        logger.info("Writing blob to local file: %s", local_path)
        local_path.write_text(blob_json, encoding="utf-8")
    else:
        logger.info("Uploading to DBFS …")
        dbfs_dir = dbfs_path.rsplit("/", 1)[0]
        subprocess.run(
            ["databricks", "fs", "mkdirs", dbfs_dir, "--profile", DATABRICKS_PROFILE],
            capture_output=True, timeout=30,
        )
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".json", delete=False, encoding="utf-8"
        ) as tmp:
            tmp.write(blob_json)
            tmp_path = tmp.name
        try:
            r = subprocess.run(
                ["databricks", "fs", "cp", tmp_path, dbfs_path,
                 "--overwrite", "--profile", DATABRICKS_PROFILE],
                capture_output=True, text=True, timeout=600,
            )
            if r.returncode != 0:
                logger.error("Upload failed:\n%s", r.stderr)
                sys.exit(1)
        finally:
            os.unlink(tmp_path)

    elapsed = time.time() - t0
    logger.info(
        "\n Done in %.0fs\n"
        "  UIF rows  : %d\n"
        "  OP rows   : %d\n"
        "  DP rows   : %d\n"
        "  SRO rows  : %d\n"
        "  FR rows   : %d\n"
        "  Blob size : %.2f MB",
        elapsed, len(uif_rows), len(op_records), len(dp_rows),
        len(sro_records), len(fr_rows), size_mb,
    )


if __name__ == "__main__":
    main()
